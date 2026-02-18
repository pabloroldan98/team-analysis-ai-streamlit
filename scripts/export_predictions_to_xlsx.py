"""
One-off script: Load all players for "today", compute predicted values, export to xlsx.

Usage:
    python -m scripts.export_predictions_to_xlsx [--output FILE.xlsx] [--verbose]

Output xlsx has two sheets:
  - Resumen: jugadores, equipo actual, precio, valor, valor/precio, valor-precio, percentil(valor)/percentil(precio)
  - Detalle: all players with all attributes as columns
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

import numpy as np

ROOT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT_DIR))

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from player import Player
from simulator.data_loader import get_active_players_at_season_start
from simulator.transfer_simulator import TransferSimulator


def _percentile_ranks_vectorized(values: np.ndarray) -> np.ndarray:
    """Compute percentile rank (0-100) for each value. O(n log n)."""
    valid = values[np.isfinite(values)]
    if len(valid) == 0:
        return np.full_like(values, np.nan)
    n = len(valid)
    sorted_vals = np.sort(valid)
    counts = np.searchsorted(sorted_vals, values, side="right")
    return np.where(np.isfinite(values), 100.0 * counts / n, np.nan)


def main():
    parser = argparse.ArgumentParser(description="Export predicted values to xlsx")
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="Output xlsx path (default: predictions_export_YYYYMMDD_HHMMSS.xlsx)",
    )
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose progress")
    args = parser.parse_args()

    if args.verbose:
        print("Loading active players for 'today'...")
    all_players = get_active_players_at_season_start("today", verbose=args.verbose)
    if args.verbose:
        print(f"  Loaded {len(all_players)} players")

    sim = TransferSimulator(club_name="Export", season="today", transfer_budget=0, salary_budget=0)
    if args.verbose:
        print("Predicting future values for all players...")
    all_players = sim._predict_values(all_players, verbose=args.verbose)

    # Filter to players with valid market_value and predicted_value
    valid = [
        p for p in all_players
        if (p.market_value is not None and p.market_value > 0
            and p.predicted_value is not None and p.predicted_value > 0)
    ]
    if args.verbose:
        print(f"  {len(valid)} players with valid precio and valor")

    precios = np.array([p.market_value for p in valid], dtype=float)
    valores = np.array([p.predicted_value for p in valid], dtype=float)
    pct_precio = _percentile_ranks_vectorized(precios)
    pct_valor = _percentile_ranks_vectorized(valores)

    # Build summary rows
    summary_rows = []
    for i, p in enumerate(valid):
        precio = p.market_value
        valor = p.predicted_value
        ratio = valor / precio if precio and precio > 0 else np.nan
        diff = valor - precio
        pp = pct_precio[i]
        pv = pct_valor[i]
        pct_ratio = pv / pp if pp and pp > 0 else np.nan
        summary_rows.append({
            "jugadores": p.name,
            "equipo_actual": p.team or "",
            "precio": precio,
            "valor": valor,
            "valor_precio": ratio,
            "valor_menos_precio": diff,
            "percentil_valor_entre_percentil_precio": pct_ratio,
        })

    # Compute percentiles for ALL players (for detail sheet)
    all_precios = np.array([p.market_value if p.market_value else np.nan for p in all_players], dtype=float)
    all_valores = np.array([p.predicted_value if p.predicted_value is not None else np.nan for p in all_players], dtype=float)
    all_pct_precio = _percentile_ranks_vectorized(all_precios)
    all_pct_valor = _percentile_ranks_vectorized(all_valores)

    # Build full-detail rows (all player attributes + summary columns at end)
    detail_col_order = [
        "player_id", "name", "team", "team_id", "position", "main_position", "other_positions",
        "age", "birth_date", "nationality", "height", "preferred_foot", "shirt_number",
        "market_value", "predicted_value", "on_loan", "loaning_team", "loaning_team_id",
        "img_url", "profile_url", "season",
        "precio", "valor", "valor/precio", "percentil(valor)/percentil(precio)", "valor-precio",
    ]
    detail_rows = []
    for i, p in enumerate(all_players):
        row = p.to_dict()
        row["predicted_value"] = p.predicted_value
        row["other_positions"] = ",".join(row.get("other_positions") or [])
        # Summary columns at end
        precio = p.market_value or np.nan
        valor = p.predicted_value if p.predicted_value is not None else np.nan
        ratio = valor / precio if precio and precio > 0 else np.nan
        diff = (valor - precio) if np.isfinite(valor) and np.isfinite(precio) else np.nan
        pp, pv = all_pct_precio[i], all_pct_valor[i]
        pct_ratio = pv / pp if pp and pp > 0 else np.nan
        row["precio"] = precio
        row["valor"] = valor
        row["valor/precio"] = ratio
        row["valor-precio"] = diff
        row["percentil(valor)/percentil(precio)"] = pct_ratio
        detail_rows.append({k: row.get(k, "") for k in detail_col_order})

    # Write xlsx
    out_path = args.output
    if not out_path:
        # ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        # out_path = ROOT_DIR / f"predictions_export_{ts}.xlsx"
        ts = datetime.now().strftime("d%m%Y%")
        out_path = ROOT_DIR / f"player_predictions-{ts}_to_one_year.xlsx"
    out_path = Path(out_path)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    summary_cols = [
        "jugadores", "equipo_actual", "precio", "valor",
        "valor/precio", "percentil(valor)/percentil(precio)", "valor-precio"
    ]
    summary_keys = [
        "jugadores", "equipo_actual", "precio", "valor",
        "valor_precio", "percentil_valor_entre_percentil_precio", "valor_menos_precio"
    ]
    ws_summary.append(summary_cols)
    for row in summary_rows:
        ws_summary.append([row[k] for k in summary_keys])

    ws_detail = wb.create_sheet("Detalle")
    ws_detail.append(detail_col_order)
    for row in detail_rows:
        ws_detail.append([row[k] for k in detail_col_order])

    # Set column widths for both sheets
    col_width = 18
    for ws in (ws_summary, ws_detail):
        for c in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(c)].width = col_width

    wb.save(out_path)
    if args.verbose:
        print(f"\nSaved to {out_path}")


if __name__ == "__main__":
    main()
